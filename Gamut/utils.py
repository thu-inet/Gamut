from Spectrum import Spectrum
from Operator import Operator
import numpy as np
from scipy.optimize import curve_fit
from re import match

def MBCTransformer(operator):

    if not issubclass(operator, Operator):
        raise TypeError('Input must be a Operator class.')
    else:
        class MBCOperator(operator):

            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._label = f'-MBC[{self._label}]'

            def __run__(self, spectrum):
                smthed = super().__run__(spectrum)
                ratio = Spectrum(smthed.counts / spectrum.counts)
                smthed_ratio = super().__run__(ratio)
                return Spectrum(smthed_ratio.counts * smthed.counts,
                                label=self._label)

    return MBCOperator

def Differential(counts, order=2, half_width=3, derive_order=1):
    '''
    Numerical differential peak-searching method based on Savitzy-Colay fitting method.
    Equation: 
        Ab + E = y
        -b = [b0, b1, ..., bi,..., bn], fitted polynomial
        -A = [[1   m     m^2    ,...,       m^n] 
              [1  m-1  (m-1)^2  ,...,   (m-1)^n]
              [1   i     i^2    ,...,       i^n]
              [1  -m   (-m)^2   ,...,    (-m)^n]], index for fitted polynomial
        -y = [y(m), y(m-1),..., y(0),..., y(-m)], windowed spectrum
        -m = half_width of transformation window
        -n = order of fitted polynomial
        -bi = polynomial coefficients
    Solution: 
        -b = (A.t*A).I * A.t * y
        -y(fitted) = Ab = A * (A.t*A).I *A.t * y = A * M * y
        -y(0)(kth derivative)(fitted) 
            = y(0)(fitted)(kth derivative)
            = ( b0 + b1 * m + ... + bn * m^n )(kth derivative)|(m=0)
            = k! * bk + (k+1)!/1! * b(k+1) * m + ... + n!/(n-k)! * bn * m^(n-k)|(m=0)
            = k! * bk
            = K! * (M * y)[kth element] 
            = k! * (M * y)[k] 
        * M = (A.t*A).I *A.t, b = M * y
    
    :param counts: spectrum counts
    :param order[2]: polymonial order of target fitted function
    :param half_width[3]: halfwidth of transform window
    :param derive_order[1]: derivative order of numerical differential
    
    :return diff: target numerical differential
    :return coefs: coefficient vector of transform 
    '''
    mat_order, mat_width = np.meshgrid(np.arange(order+1), np.arange(-half_width, half_width+1))
    A = mat_width ** mat_order
    M = np.matmul( np.linalg.inv(np.matmul(A.T, A)), A.T )
    diff = np.zeros(counts.shape)
    if derive_order == 0:
        for i in range(half_width, counts.shape[0]-half_width):
            diff[i] = np.matmul( M, counts[i-half_width: i+half_width+1] )[derive_order]
        coefs = M[derive_order]
    else:
        for i in range(half_width, counts.shape[0]-half_width):
            diff[i] = np.matmul( M, counts[i-half_width: i+half_width+1] )[derive_order] * derive_order
        coefs = M[derive_order] * derive_order
    return diff, coefs


# def Peakfit(counts, npeak):
#     index = np.arange(counts.shape[0])
#     low = np.zeros(3*npeak+1)
#     high = np.zeros(3*npeak+1)
#     for i in range(npeak):
#         low[i*3] = np.min(counts)
#         high[i*3] = np.max(counts)
#         low[i*3+1] = 0
#         high[i*3+1] = counts.shape[0]
#         low[i*3+2] = 1
#         high[i*3+2] = 7
#     low[-1] = min(counts)
#     high[-1] = min(counts)+50

#     vars = (low+high)/2
#     for i in range(npeak):
#         vars[i*3+1] = (high[i*3+1]-low[i*3+1]) * (i+1)/(npeak+2) + low[i*3+1]

#     vars, covs = curve_fit(fitfunc, index, counts, p0=vars, bounds=(low, high))

#     return vars

def gauss(index, amplitude, centroid, stderror):
    return amplitude * np.exp(-(index-centroid)**2 / 2 / stderror**2)


def compound(index, *args):
    counts = args[-1]
    for i in range(len(args)//3):
        counts += gauss(index, *args[i*3: i*3+3])
    return counts

def compound(index, amplitude, centroid, stderror):
    return amplitude * np.exp(-(index-centroid)**2 / 2 / stderror**2)

def peak_fit(counts, npeak=1):
    index = np.arange(counts.shape[0])
    try:
        vars = curve_fit(compound, index, counts, p0=np.ones(3*npeak))[0]
    except:
        vars = np.ones(3*npeak)
    return vars

class Nuclide():

    def __init__(self, symbol, mass, half_life):
        self._symbol = symbol
        self._mass = mass
        self._half_life = half_life

    def add_gamma(self, energy, branch):
        if not hasattr(self, '_gamma_branch'):
            self._gamma_branch = {}
        self._gamma_branch[energy] = branch

def match_unit(unit):
    match unit:
        case 'Hrs.':
            unit = 3600
        case 'Min.':
            unit = 60
        case 'Sec.':
            unit = 1
        case 'Yrs.':
            unit = 3600*24*365
        case 'Days':
            unit = 3600*24
    return unit

def convert_gv_report(path):
    with open(path, 'r') as fileopen:
        filelines = fileopen.readlines()
    index = 0
    list_nuc = []
    while index < len(filelines):
        line = filelines[index]
        subindex = 1
        if (m := match(r"\s+([\w]+)-([\d]+)\s+([\d\.]+)\s+([\w\.]+)\s+([\d\.]+)\s+([-\w]+)", line)):
            nuc = Nuclide(symbol=m.group(1), mass=int(m.group(2)), half_life=float(m.group(3))*match_unit(m.group(4)))
            while (m := match(r"\s+([\d.]+)keV\s+([\d.]+)%\s+([-\w]+)", filelines[index+subindex])):
                nuc.add_gamma(energy=float(m.group(1)), branch=float(m.group(2))/100)
                subindex += 1
            list_nuc.append(nuc)
        index += subindex

    return list_nuc

if __name__ == '__main__':

    # l = np.array([555.73042307, 663.71116222, 710.3171541 , 673.10288198,       566.60760154])
    # vars = peak_fit(l, 1)
    # l2 = compound(np.arange(l.shape[0]), *vars)

    # import matplotlib.pyplot as plt
    # plt.plot(l)
    # plt.plot(l2)
    # plt.plot()
    # plt.show()

    list_nuc = convert_gv_report('./Suspect.Txt')
    import openpyxl as xl
    wb = xl.Workbook()
    ws = wb.active
    index = 1
    for nuc in list_nuc:
        subindex = 1
        ws['A'+str(index)] = nuc._symbol + str(nuc._mass)
        ws['B'+str(index)] = nuc._half_life
        # if hasattr(nuc, '_gamma_branch'):
        #     for key, value in nuc._gamma_branch.items():
        #         ws['B'+str(index+subindex)] = key
        #         ws['C'+str(index+subindex)] = value
        #         subindex += 1
        index += subindex
    wb.save('./Suspect.xlsx')

    
    
    

    # with open('./Suspect.Txt', 'r') as fileopen:
    #     filelines = fileopen.readlines()

    # line = filelines[4]
    # print(line)
    # print(match(r"\s+(\w+)-(\d+)\s+([\d\.]+)\s+([\w\.]+)\s+([\d\.]+)\s+([-\w]+)").group())
